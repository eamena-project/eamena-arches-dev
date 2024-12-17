# record the "APAAME Master Catalog" (AMC) file/folder hierarchical structure (FHS) in an XLSX file

#%% 
# variables
import os
from openpyxl import Workbook

# in file
AMC = "D:/APAAME Master Catalog"
# AMC = "C:/Rprojects/eamena-arches-dev/projects"
AMC = "D:/TEMP"
# out file
FHS = "C:/Rprojects/eamena-arches-dev/projects/apaame/temp-amc-fhs-1.xlsx"

l = [x[0] for x in os.walk(AMC)]

wb = Workbook()
ws = wb.active
ws.title = "APAAME Master Catalog FHS"

n = 0
#%% 
# run

# for row_num, item in enumerate(l, start=1):
# 	n = n + 1
# 	columns = item.split('\\')
# 	ws.append(columns)
# 	if (n % 100) == 0:
# 		print(" read record " + str(n))
# for row_num, item in enumerate(os.walk(AMC), start=1):
#     # Only proceed if the current path is within the first two levels of subfolders
#     depth = len(item[0].replace(AMC, '').split(os.sep)) - 1
#     if depth <= 2:
#         n += 1
#         columns = item[0].split(os.sep)[-depth:]  # Slice to keep only relevant levels
#         print(columns)
#         ws.append(columns)
#         if n % 100 == 0:
#             print("read record " + str(n))
for root, dirs, files in os.walk(AMC):
    # Calculate the depth of the current directory
    depth = len(root.replace(AMC, '').strip(os.sep).split(os.sep))
    # Include only the base directory and two levels of subfolders
    if depth <= 2:
        n += 1
        # Append the full path of the current directory to the worksheet
        ws.append([root])
        if n % 100 == 0:
            print(f"read record {n}")
      
wb.save(FHS)
print(os.path.basename(FHS) + " has been exported")
# %%


import os
from openpyxl import Workbook

# Input and output file paths
AMC = "D:/APAAME Master Catalog"
# AMC = "C:/Rprojects/eamena-arches-dev/projects"
AMC = "D:/TEMP"
# out file
FHS = "C:/Rprojects/eamena-arches-dev/projects/apaame/temp-amc-fhs-1.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "APAAME Master Catalog FHS"

n = 0
for root, dirs, files in os.walk(AMC):
    # Calculate the depth of the current directory
    depth = len(root.replace(AMC, '').strip(os.sep).split(os.sep))
    # Include only the base directory and two levels of subfolders
    if depth <= 2:
        n += 1
        # Append the full path of the current directory to the worksheet
        ws.append([root])
        if n % 100 == 0:
            print(f"read record {n}")

# Save the workbook
wb.save(FHS)
print(f"{os.path.basename(FHS)} has been exported")


# %%

import os
from openpyxl import Workbook

# Input and output file paths
# Input and output file paths
AMC = "D:/APAAME Master Catalog"
# AMC = "C:/Rprojects/eamena-arches-dev/projects"
AMC = "D:/TEMP"
# out file
FHS = "C:/Rprojects/eamena-arches-dev/projects/apaame/temp-amc-fhs-1.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "APAAME Master Catalog FHS"

# We'll use a variable to keep track of the last level 2 folder added to column A
last_level_2_path = None

n = 0
for root, dirs, files in os.walk(AMC):
    # Calculate the depth of the current directory
    depth = len(root.replace(AMC, '').strip(os.sep).split(os.sep))
    
    # Handle paths up to level 2 (for column A)
    if depth <= 2:
        ws.append([root])
        last_level_2_path = root
        n += 1
    elif depth == 3:
        # For level 3 subfolders, find the row of the last level 2 folder in column A and append this path to column B
        if last_level_2_path:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
                if row[0].value == last_level_2_path:
                    ws.cell(row=row[0].row, column=2, value=root)
                    break
        n += 1

    if n % 100 == 0:
        print(f"read record {n}")

# Save the workbook
wb.save(FHS)
print(f"{os.path.basename(FHS)} has been exported")

# %%

n = 0
for root, dirs, files in os.walk(AMC):
    # Calculate the depth of the current directory
    depth = len(root.replace(AMC, '').strip(os.sep).split(os.sep))
    
    # Handle up to level 2 folders (including base folder)
    if depth <= 2:
        # Append level 1 and level 2 folders directly to column A
        ws.append([root, ""])
    elif depth == 3:
        # For level 3 subfolders, append to the next row in column B, but also copy its parent path to column A for clarity
        parent_path = os.path.dirname(root)  # Get the parent path of the level 3 subfolder
        ws.append([parent_path, root])
    # Skip folders of level 4 and beyond

wb.save(FHS)
print(f"{os.path.basename(FHS)} has been exported")

# %%

wb = Workbook()
ws = wb.active
ws.title = "APAAME Master Catalog FHS"

for root, dirs, files in os.walk(AMC):
    # Calculate the depth of the current directory
    depth = len(root.replace(AMC, '').strip(os.sep).split(os.sep))
    
    # Handle up to level 2 folders (including base folder)
    if depth <= 2:
        # Append level 1 and level 2 folders directly to column A
        ws.append([root, ""])
    elif depth == 3:
        # For level 3 subfolders, append to the next row in column B, but also copy its parent path to column A for clarity
        parent_path = os.path.dirname(root)  # Get the parent path of the level 3 subfolder
        ws.append([parent_path, root])
    # Skip folders of level 4 and beyond

wb.save(FHS)
print(f"{os.path.basename(FHS)} has been exported")
# %%

wb = Workbook()
ws = wb.active
ws.title = "APAAME Master Catalog FHS"

for root, dirs, files in os.walk(AMC):
    # Calculate the depth of the current directory relative to AMC
    depth = len(root.replace(AMC, '').strip(os.sep).split(os.sep))
    
    # Process only directories at level 1 or 2
    if depth == 1:
        # Level 1 subfolders are added to column A
        ws.append([os.path.basename(root)])
    elif depth == 2:
        # Level 2 subfolders are added to column B
        # Find the row of the last entry in column A to append the subfolder in column B
        last_row = ws.max_row
        parent_folder = os.path.basename(os.path.dirname(root))
        subfolder_name = os.path.basename(root)
        
        # Check if the parent folder of the current subfolder is already listed in the last row
        if ws.cell(row=last_row, column=1).value == parent_folder:
            ws.cell(row=last_row, column=2, value=subfolder_name)
        else:
            # If the parent folder is different, it means we've moved to a new set of subfolders
            ws.append([parent_folder, subfolder_name])
    # Skip level 3 and beyond
    else:
        continue

# Save the workbook
wb.save(FHS)
print(f"{os.path.basename(FHS)} has been exported")
# %%

import os
from openpyxl import Workbook

# # Input and output file paths
# AMC = "D:/TEMP"  # This is the base directory you're working with
# FHS = "C:/path/to/your/output.xlsx"  # Output file where the Excel will be saved

wb = Workbook()
ws = wb.active
ws.title = "APAAME Master Catalog FHS"

for root, dirs, files in os.walk(AMC):
    # Calculate the depth of the current directory
    depth = len(root.replace(AMC, '').strip(os.sep).split(os.sep))
    text = "- [ ] " + root 
    # Level 1 directories: full folder names in column A
    if depth == 1:
        ws.append([text, ""])  # Add the full path of level 1 folders to column A, leave column B empty
    # Level 2 directories: full subfolder names in column B
    elif depth == 2:
        ws.append(["", text])  # Leave column A empty, add the full path of level 2 folders to column B

# No processing for level 3 or beyond, as requested

# Save the workbook
wb.save(FHS)
print(f"{os.path.basename(FHS)} has been exported")

# %%


